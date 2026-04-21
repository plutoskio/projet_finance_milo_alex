from __future__ import annotations

from html import escape
from pathlib import Path
import re

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUT = ROOT / "exports"
OUT.mkdir(exist_ok=True)
MODEL_XLSX = ROOT / "excel_model" / "Michelin_valuation_model.xlsx"
LEGACY_REMOVED_CHARTS = [
    "drivers_risks_heatmap.svg",
    "peer_screening_matrix.svg",
    "wacc_build_up.svg",
    "investment_scorecard.svg",
]

BLUE = "#1F4E5F"
TEAL = "#2A9D8F"
RED = "#D1495B"
GOLD = "#F4A261"
LIGHT = "#F6F8FA"
MID = "#E6EBEF"
DARK = "#20252B"
GRAY = "#6B7280"


def save_svg(name: str, content: str) -> None:
    (OUT / name).write_text(content, encoding="utf-8")


def header(title: str, subtitle: str, width: int = 1050, height: int = 620) -> str:
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">
  <rect width="{width}" height="{height}" fill="#ffffff"/>
  <text x="70" y="38" font-family="Arial, Helvetica, sans-serif" font-size="22" font-weight="700" fill="{BLUE}">{escape(title)}</text>
  <text x="70" y="60" font-family="Arial, Helvetica, sans-serif" font-size="12" fill="{GRAY}">{escape(subtitle)}</text>
'''


def footer(source: str, y: int = 590) -> str:
    return f'  <text x="70" y="{y}" font-family="Arial, Helvetica, sans-serif" font-size="10" fill="{GRAY}">{escape(source)}</text>\n</svg>\n'


def scenario_model_values() -> dict:
    wb = load_workbook(MODEL_XLSX, data_only=False)
    hist = wb["Historical"]
    forecast = wb["Forecast"]
    wacc_ws = wb["WACC"]
    sensitivity = wb["Sensitivity"]

    hist_sales = float(hist["G4"].value)
    share_price = float(wacc_ws["B5"].value)
    shares = float(wacc_ws["B6"].value)
    net_debt = float(hist["G18"].value)
    risk_free = float(wacc_ws["B10"].value)
    erp = float(wacc_ws["B11"].value)
    beta = float(wacc_ws["B12"].value)
    pre_tax_cost_debt = float(wacc_ws["B14"].value)
    tax_rate = float(wacc_ws["B15"].value)
    market_cap = share_price * shares
    enterprise_value_market = market_cap + net_debt
    base_wacc = ((risk_free + beta * erp) * (market_cap / enterprise_value_market)) + (
        pre_tax_cost_debt * (1 - tax_rate) * (net_debt / enterprise_value_market)
    )
    lookup = {
        "Forecast!F8": float(forecast["F8"].value),
        "Forecast!F14": float(forecast["F14"].value),
        "WACC!B19": base_wacc,
        "WACC!B20": float(wacc_ws["B20"].value),
    }

    def resolve_simple_formula(value):
        if isinstance(value, (int, float)):
            return float(value)
        if not isinstance(value, str):
            raise ValueError(f"Unsupported scenario value: {value!r}")
        match = re.fullmatch(r"=(Forecast!F8|Forecast!F14|WACC!B19|WACC!B20)([+-]\d+(?:\.\d+)?)?", value)
        if not match:
            raise ValueError(f"Unsupported scenario formula: {value}")
        base = lookup[match.group(1)]
        offset = float(match.group(2) or 0.0)
        return base + offset

    growth = [float(forecast[cell].value) for cell in ["B5", "C5", "D5", "E5", "F5"]]
    dna_rates = [float(forecast[cell].value) for cell in ["B12", "C12", "D12", "E12", "F12"]]
    base_margin_2026 = float(forecast["B8"].value)
    base_capex_2026 = float(forecast["B14"].value)

    sales = []
    prev_sales = hist_sales
    for g in growth:
        prev_sales = prev_sales * (1 + g)
        sales.append(prev_sales)

    scenarios = []
    for row, name in zip([5, 6, 7], ["Bear", "Base", "Bull"]):
        margin_terminal = resolve_simple_formula(sensitivity[f"B{row}"].value)
        capex_terminal = resolve_simple_formula(sensitivity[f"C{row}"].value)
        scenario_wacc = resolve_simple_formula(sensitivity[f"D{row}"].value)
        terminal_growth = resolve_simple_formula(sensitivity[f"E{row}"].value)

        margins = [base_margin_2026] + [
            base_margin_2026 + (margin_terminal - base_margin_2026) * i / 4 for i in range(1, 5)
        ]
        capex_rates = [base_capex_2026] + [
            base_capex_2026 + (capex_terminal - base_capex_2026) * i / 4 for i in range(1, 5)
        ]

        fcff = []
        prev_sales = hist_sales
        for idx, year_sales in enumerate(sales):
            ebit = year_sales * margins[idx]
            nopat = ebit * (1 - tax_rate)
            d_and_a = year_sales * dna_rates[idx]
            capex = year_sales * capex_rates[idx]
            delta_bfr = (year_sales - prev_sales) * 0.01
            fcff.append(nopat + d_and_a - capex - delta_bfr)
            prev_sales = year_sales

        discount_factors = [1 / ((1 + scenario_wacc) ** (i + 1)) for i in range(5)]
        pv_fcff = sum(f * d for f, d in zip(fcff, discount_factors))
        terminal_value = fcff[-1] * (1 + terminal_growth) / (scenario_wacc - terminal_growth)
        enterprise_value = pv_fcff + terminal_value * discount_factors[-1]
        equity_value = enterprise_value - net_debt
        price = equity_value / shares
        upside = price / share_price - 1

        scenarios.append(
            {
                "name": name,
                "price": price,
                "upside": upside,
            }
        )

    return {"share_price": share_price, "shares": shares, "scenarios": scenarios}


def scenario_dcf_values() -> str:
    payload = scenario_model_values()
    current_price = payload["share_price"]
    scenarios = payload["scenarios"]
    width, height = 1000, 560
    x0, y0, w, h = 120, 90, 780, 340
    colors = {"Bear": RED, "Base": TEAL, "Bull": BLUE}
    max_val = max(65.0, max(item["price"] for item in scenarios) + 7.0)

    def yp(v: float) -> float:
        return y0 + h - (v / max_val) * h

    s = header(
        "Scenario DCF - value per share",
        "Growth stays fixed at Forecast. Bear / Base / Bull vary terminal margin, Capex/CA, WACC and terminal growth.",
        width=width,
        height=height,
    )
    s += f'\n  <line x1="{x0}" y1="{y0+h}" x2="{x0+w}" y2="{y0+h}" stroke="{DARK}" stroke-width="1.2"/>\n'
    s += f'  <line x1="{x0}" y1="{y0}" x2="{x0}" y2="{y0+h}" stroke="{DARK}" stroke-width="1.2"/>\n'
    s += '  <g font-family="Arial, Helvetica, sans-serif" font-size="11" fill="#6B7280">\n'
    tick = 0
    while tick <= int(max_val):
        y = yp(tick)
        s += f'    <line x1="{x0}" y1="{y:.1f}" x2="{x0+w}" y2="{y:.1f}" stroke="{MID}" stroke-width="1"/>\n'
        s += f'    <text x="{76 if tick >= 10 else 82}" y="{y+4:.1f}">{tick}</text>\n'
        tick += 10
    s += f'    <text x="63" y="{y0+5}">EUR/share</text>\n'
    s += "  </g>\n"

    y_price = yp(current_price)
    s += f'  <line x1="{x0}" y1="{y_price:.1f}" x2="{x0+w}" y2="{y_price:.1f}" stroke="{RED}" stroke-width="2" stroke-dasharray="6 4"/>\n'
    s += f'  <text x="{x0+w+5}" y="{y_price+3:.1f}" font-family="Arial, Helvetica, sans-serif" font-size="12" fill="{RED}">Current price EUR {current_price:.1f}</text>\n'

    bar_positions = [235, 440, 645]
    for item, x in zip(scenarios, bar_positions):
        price = item["price"]
        y = yp(price)
        bh = y0 + h - y
        label = f"EUR {price:.1f}"
        upside = f"{item['upside']:+.1%} upside"
        s += f'  <rect x="{x}" y="{y:.1f}" width="120" height="{bh:.1f}" fill="{colors[item["name"]]}"/>\n'
        s += f'  <text x="{x+60}" y="{y-12:.1f}" text-anchor="middle" font-family="Arial, Helvetica, sans-serif" font-size="18" font-weight="700" fill="{DARK}">{label}</text>\n'
        s += f'  <text x="{x+60}" y="462" text-anchor="middle" font-family="Arial, Helvetica, sans-serif" font-size="14" font-weight="700" fill="{DARK}">{item["name"]}</text>\n'
        s += f'  <text x="{x+60}" y="483" text-anchor="middle" font-family="Arial, Helvetica, sans-serif" font-size="11" fill="{GRAY}">{upside}</text>\n'

    s += footer("Source: Michelin valuation model workbook, Sensitivity sheet.", y=535)
    return s


def buyback_model_values() -> dict:
    payload = scenario_model_values()
    share_price = payload["share_price"]
    shares = payload["shares"]
    base_price = next(item["price"] for item in payload["scenarios"] if item["name"] == "Base")
    program = 2000.0
    repurchased = program / share_price
    post_shares = shares - repurchased
    share_reduction = repurchased / shares
    eps_accretion = shares / post_shares - 1
    post_price = ((base_price * shares) - program) / post_shares
    dcf_accretion = post_price / base_price - 1
    return {
        "share_price": share_price,
        "shares": shares,
        "program": program,
        "base_price": base_price,
        "repurchased": repurchased,
        "post_shares": post_shares,
        "share_reduction": share_reduction,
        "eps_accretion": eps_accretion,
        "post_price": post_price,
        "dcf_accretion": dcf_accretion,
    }


def buyback_impact() -> str:
    data = buyback_model_values()
    width, height = 1120, 640
    s = header(
        "Illustrative EUR 2bn buyback impact",
        f"At EUR {data['share_price']:.2f}/share, EUR 2.0bn of repurchases would reduce share count by about {data['share_reduction']:.1%}.",
        width=width,
        height=height,
    )

    # Left section: share count bridge
    s += f'  <text x="95" y="108" font-family="Arial" font-size="16" font-weight="700" fill="{BLUE}">Share count bridge</text>\n'
    left_x, bar_x, max_w = 95, 270, 430
    top_y = 170
    current_shares = data["shares"]
    repurchased = data["repurchased"]
    post_shares = data["post_shares"]
    max_shares = current_shares

    rows = [
        ("Current shares", current_shares, BLUE, f"{current_shares:.1f}m"),
        ("Repurchased", repurchased, RED, f"-{repurchased:.1f}m"),
        ("Pro forma shares", post_shares, TEAL, f"{post_shares:.1f}m"),
    ]
    for i, (label, value, color, value_label) in enumerate(rows):
        y = top_y + i * 95
        bar_w = max_w * value / max_shares
        s += f'  <text x="{left_x}" y="{y+20}" font-family="Arial" font-size="13" font-weight="700" fill="{DARK}">{escape(label)}</text>\n'
        s += f'  <rect x="{bar_x}" y="{y}" width="{bar_w:.1f}" height="34" fill="{color}" rx="3"/>\n'
        if label == "Repurchased":
            s += f'  <text x="{bar_x + bar_w + 14:.1f}" y="{y+22}" font-family="Arial" font-size="13" font-weight="700" fill="{color}">{value_label}</text>\n'
        else:
            s += f'  <text x="{bar_x + bar_w + 16:.1f}" y="{y+22}" font-family="Arial" font-size="13" font-weight="700" fill="{color}">{value_label}</text>\n'

    s += f'  <rect x="95" y="438" width="645" height="62" fill="{LIGHT}" stroke="{MID}"/>\n'
    s += f'  <text x="115" y="465" font-family="Arial" font-size="13" fill="{DARK}">Share count falls by {data["share_reduction"]:.1%}; with constant net income, EPS rises by about {data["eps_accretion"]:.1%}.</text>\n'
    s += f'  <text x="115" y="487" font-family="Arial" font-size="13" fill="{DARK}">This is optical accretion first; value creation depends on the buyback price versus intrinsic value.</text>\n'

    # Right section: value per share read-across
    right_x = 790
    s += f'  <text x="{right_x}" y="108" font-family="Arial" font-size="16" font-weight="700" fill="{BLUE}">Value per share read-across</text>\n'

    card_w, card_h = 255, 84
    card_y1 = 145
    cards = [
        ("DCF / share before buyback", data["base_price"], BLUE, card_y1),
        ("DCF / share after cash spend", data["post_price"], TEAL, card_y1 + 108),
    ]
    for label, value, color, y in cards:
        s += f'  <rect x="{right_x}" y="{y}" width="{card_w}" height="{card_h}" fill="{LIGHT}" stroke="{MID}"/>\n'
        s += f'  <text x="{right_x+22}" y="{y+28}" font-family="Arial" font-size="12" fill="{GRAY}">{escape(label)}</text>\n'
        s += f'  <text x="{right_x+22}" y="{y+60}" font-family="Arial" font-size="28" font-weight="700" fill="{color}">EUR {value:.1f}</text>\n'

    s += f'  <line x1="{right_x+80}" y1="385" x2="{right_x+210}" y2="385" stroke="{TEAL}" stroke-width="5"/>\n'
    s += f'  <polygon points="{right_x+210},365 {right_x+210},405 {right_x+255},385" fill="{TEAL}"/>\n'
    s += f'  <text x="{right_x+115}" y="438" text-anchor="middle" font-family="Arial" font-size="26" font-weight="700" fill="{TEAL}">+{data["dcf_accretion"]:.1%}</text>\n'
    s += f'  <text x="{right_x+115}" y="462" text-anchor="middle" font-family="Arial" font-size="12" fill="{DARK}">Illustrative DCF/share accretion</text>\n'

    s += f'  <rect x="{right_x}" y="500" width="{card_w}" height="82" fill="#ffffff" stroke="{MID}"/>\n'
    s += f'  <text x="{right_x+22}" y="530" font-family="Arial" font-size="12" font-weight="700" fill="{DARK}">Key condition</text>\n'
    s += f'  <text x="{right_x+22}" y="554" font-family="Arial" font-size="12" fill="{DARK}">Buybacks create value only if shares</text>\n'
    s += f'  <text x="{right_x+22}" y="575" font-family="Arial" font-size="12" fill="{DARK}">are repurchased below intrinsic value.</text>\n'

    s += f'  <text x="70" y="606" font-family="Arial" font-size="13" fill="{DARK}">Readout: the buyback can reinforce the Buy case, but the core thesis still needs margin recovery and strong FCF.</text>\n'
    s += footer(
        f"Source: Michelin valuation model workbook. Illustrative calculation at EUR {data['share_price']:.2f}/share and {data['shares']:.2f}m shares.",
        y=628,
    )
    return s


def tornado_model_values() -> dict:
    wb = load_workbook(MODEL_XLSX, data_only=False)
    hist = wb["Historical"]
    forecast = wb["Forecast"]
    wacc_ws = wb["WACC"]
    sensitivity = wb["Sensitivity"]

    hist_sales = float(hist["G4"].value)
    share_price = float(wacc_ws["B5"].value)
    shares = float(wacc_ws["B6"].value)
    net_debt = float(hist["G18"].value)
    risk_free = float(wacc_ws["B10"].value)
    erp = float(wacc_ws["B11"].value)
    beta = float(wacc_ws["B12"].value)
    pre_tax_cost_debt = float(wacc_ws["B14"].value)
    tax_rate = float(wacc_ws["B15"].value)
    market_cap = share_price * shares
    enterprise_value_market = market_cap + net_debt
    base_wacc = ((risk_free + beta * erp) * (market_cap / enterprise_value_market)) + (
        pre_tax_cost_debt * (1 - tax_rate) * (net_debt / enterprise_value_market)
    )
    growth = [float(forecast[cell].value) for cell in ["B5", "C5", "D5", "E5", "F5"]]
    dna_rates = [float(forecast[cell].value) for cell in ["B12", "C12", "D12", "E12", "F12"]]
    base_margin_2026 = float(forecast["B8"].value)
    base_margin_terminal = float(forecast["F8"].value)
    base_capex_2026 = float(forecast["B14"].value)
    base_capex_terminal = float(forecast["F14"].value)
    base_g = float(wacc_ws["B20"].value)

    def price(
        margin_terminal: float = base_margin_terminal,
        capex_terminal: float = base_capex_terminal,
        wacc: float = base_wacc,
        terminal_growth: float = base_g,
    ) -> float:
        sales = []
        prev_sales = hist_sales
        for gr in growth:
            prev_sales = prev_sales * (1 + gr)
            sales.append(prev_sales)
        margins = [base_margin_2026] + [
            base_margin_2026 + (margin_terminal - base_margin_2026) * i / 4 for i in range(1, 5)
        ]
        capex_rates = [base_capex_2026] + [
            base_capex_2026 + (capex_terminal - base_capex_2026) * i / 4 for i in range(1, 5)
        ]
        fcff = []
        prev_sales = hist_sales
        for idx, year_sales in enumerate(sales):
            ebit = year_sales * margins[idx]
            nopat = ebit * (1 - tax_rate)
            d_and_a = year_sales * dna_rates[idx]
            capex = year_sales * capex_rates[idx]
            delta_bfr = (year_sales - prev_sales) * 0.01
            fcff.append(nopat + d_and_a - capex - delta_bfr)
            prev_sales = year_sales
        discount_factors = [1 / ((1 + wacc) ** (i + 1)) for i in range(5)]
        pv_fcff = sum(f * d for f, d in zip(fcff, discount_factors))
        terminal_value = fcff[-1] * (1 + terminal_growth) / (wacc - terminal_growth)
        enterprise_value = pv_fcff + terminal_value * discount_factors[-1]
        equity_value = enterprise_value - net_debt
        return equity_value / shares

    base_price = price()
    rows = [
        {
            "label": "Marge EBIT terminale 2030",
            "low_label": "10.5%",
            "high_label": "13.0%",
            "low_value": price(margin_terminal=float(forecast["F8"].value) - 0.015),
            "high_value": price(margin_terminal=float(forecast["F8"].value) + 0.010),
        },
        {
            "label": "WACC",
            "low_label": f"{(base_wacc + 0.005):.1%}",
            "high_label": f"{(base_wacc - 0.005):.1%}",
            "low_value": price(wacc=base_wacc + 0.005),
            "high_value": price(wacc=base_wacc - 0.005),
        },
        {
            "label": "Croissance terminale",
            "low_label": "1.0%",
            "high_label": "2.0%",
            "low_value": price(terminal_growth=0.010),
            "high_value": price(terminal_growth=0.020),
        },
        {
            "label": "Capex / CA terminal 2030",
            "low_label": "7.5%",
            "high_label": "6.5%",
            "low_value": price(capex_terminal=float(forecast["F14"].value) + 0.005),
            "high_value": price(capex_terminal=float(forecast["F14"].value) - 0.005),
        },
    ]
    for row in rows:
        row["amplitude"] = abs(row["high_value"] - row["low_value"])
    rows.sort(key=lambda x: x["amplitude"], reverse=True)
    return {"base_price": base_price, "rows": rows}


def tornado_sensitivity() -> str:
    payload = tornado_model_values()
    base_price = payload["base_price"]
    rows = payload["rows"]
    width, height = 1050, 560
    x0, y0, w = 290, 115, 620
    minv = min(min(r["low_value"], r["high_value"]) for r in rows) - 1.5
    maxv = max(max(r["low_value"], r["high_value"]) for r in rows) + 1.5

    def xp(v: float) -> float:
        return x0 + (v - minv) / (maxv - minv) * w

    s = header(
        "Tornado sensitivity - DCF value per share",
        "One variable changes at a time; revenue growth stays fixed at Forecast.",
        width=width,
        height=height,
    )
    tick = round(minv)
    while tick <= round(maxv):
        x = xp(tick)
        s += f'  <line x1="{x:.1f}" y1="95" x2="{x:.1f}" y2="430" stroke="{MID}"/>\n'
        s += f'  <text x="{x-8:.1f}" y="455" font-family="Arial" font-size="11" fill="{GRAY}">{tick:.0f}</text>\n'
        tick += 2
    s += f'  <line x1="{xp(base_price):.1f}" y1="95" x2="{xp(base_price):.1f}" y2="430" stroke="{DARK}" stroke-width="2" stroke-dasharray="6 4"/>\n'
    s += f'  <text x="{xp(base_price)+8:.1f}" y="108" font-family="Arial" font-size="12" fill="{DARK}">Base: EUR {base_price:.1f}</text>\n'

    for i, row in enumerate(rows):
        y = y0 + i * 78
        low, high = row["low_value"], row["high_value"]
        lo_x, hi_x = xp(low), xp(high)
        mid_x = xp(base_price)
        s += f'  <text x="90" y="{y+8}" font-family="Arial" font-size="13" font-weight="700" fill="{DARK}">{escape(row["label"])}</text>\n'
        s += f'  <rect x="{min(lo_x, mid_x):.1f}" y="{y-10}" width="{abs(mid_x-lo_x):.1f}" height="22" fill="{RED}" opacity="0.85"/>\n'
        s += f'  <rect x="{min(mid_x, hi_x):.1f}" y="{y-10}" width="{abs(hi_x-mid_x):.1f}" height="22" fill="{TEAL}" opacity="0.90"/>\n'
        s += f'  <text x="{lo_x-6:.1f}" y="{y-16}" text-anchor="end" font-family="Arial" font-size="11" fill="{GRAY}">{row["low_label"]}</text>\n'
        s += f'  <text x="{hi_x+6:.1f}" y="{y-16}" font-family="Arial" font-size="11" fill="{GRAY}">{row["high_label"]}</text>\n'
        s += f'  <text x="{lo_x-6:.1f}" y="{y+7}" text-anchor="end" font-family="Arial" font-size="11" fill="{DARK}">{low:.1f}</text>\n'
        s += f'  <text x="{hi_x+6:.1f}" y="{y+7}" font-family="Arial" font-size="11" fill="{DARK}">{high:.1f}</text>\n'

    s += f'  <text x="90" y="515" font-family="Arial" font-size="13" fill="{DARK}">Readout: margin is the biggest driver; WACC is second. Capex and terminal growth matter, but less than operating margin restoration.</text>\n'
    s += footer("Source: Michelin valuation model workbook, Tornado logic from Base vs one-variable changes.", y=535)
    return s


def market_segmentation() -> str:
    data = [
        ("PLT remplacement", 1255.6, TEAL),
        ("PLT OE", 418.7, BLUE),
        ("Truck remplacement", 171.6, GOLD),
        ("Truck OE", 47.2, RED),
    ]
    total = sum(v for _, v, _ in data)
    x0, y0, w, h = 120, 155, 800, 64
    s = header(
        "2025 tire market volumes by segment",
        "Global volume in millions of tires. Replacement is structurally larger and more resilient than OE.",
    )
    cur = x0
    for label, val, color in data:
        bw = w * val / total
        s += f'  <rect x="{cur:.1f}" y="{y0}" width="{bw:.1f}" height="{h}" fill="{color}"/>\n'
        if bw > 65:
            s += f'  <text x="{cur + bw/2:.1f}" y="{y0 + 38}" text-anchor="middle" font-family="Arial" font-size="13" font-weight="700" fill="#ffffff">{val:,.0f}</text>\n'
        cur += bw
    s += f'  <rect x="{x0}" y="{y0}" width="{w}" height="{h}" fill="none" stroke="{DARK}" stroke-width="1"/>\n'
    s += '  <g font-family="Arial, Helvetica, sans-serif" font-size="13" fill="#20252B">\n'
    lx, ly = 120, 280
    for i, (label, val, color) in enumerate(data):
        y = ly + i * 45
        pct = val / total
        s += f'    <rect x="{lx}" y="{y-14}" width="16" height="16" fill="{color}"/>\n'
        s += f'    <text x="{lx+26}" y="{y}">{escape(label)}: {val:,.1f}m tires ({pct:.1%})</text>\n'
    s += "  </g>\n"
    s += f'''  <text x="120" y="505" font-family="Arial" font-size="15" font-weight="700" fill="{BLUE}">Key read-across</text>
  <text x="120" y="532" font-family="Arial" font-size="13" fill="{DARK}">The replacement market is much larger than OE: this supports Michelin's resilience because tire replacement depends on the installed vehicle base, not only new vehicle production.</text>
'''
    s += footer("Source: Michelin key figures 2020-2025, Markets tab.")
    return s


def competitive_positioning() -> str:
    points = [
        ("Michelin", 82, 82, RED),
        ("Bridgestone", 72, 68, TEAL),
        ("Pirelli", 86, 45, TEAL),
        ("Continental", 58, 62, TEAL),
        ("Yokohama", 48, 43, TEAL),
        ("Hankook", 42, 38, TEAL),
        ("Goodyear", 35, 30, GRAY),
    ]
    x0, y0, w, h = 120, 100, 760, 400
    def xp(v): return x0 + w * v / 100
    def yp(v): return y0 + h * (1 - v / 100)
    s = header(
        "Competitive positioning map",
        "Qualitative map: premium positioning vs technical/specialty exposure. Michelin is differentiated by brand, R&D and specialties.",
    )
    s += f'  <rect x="{x0}" y="{y0}" width="{w}" height="{h}" fill="{LIGHT}" stroke="{DARK}"/>\n'
    s += f'  <line x1="{x0+w/2}" y1="{y0}" x2="{x0+w/2}" y2="{y0+h}" stroke="{MID}" stroke-width="2"/>\n'
    s += f'  <line x1="{x0}" y1="{y0+h/2}" x2="{x0+w}" y2="{y0+h/2}" stroke="{MID}" stroke-width="2"/>\n'
    s += f'  <text x="{x0+15}" y="{y0+25}" font-family="Arial" font-size="12" fill="{GRAY}">More technical / specialties</text>\n'
    s += f'  <text x="{x0+w-210}" y="{y0+h-15}" font-family="Arial" font-size="12" fill="{GRAY}">More premium positioning</text>\n'
    s += f'  <text x="{x0+15}" y="{y0+h-15}" font-family="Arial" font-size="12" fill="{GRAY}">More commodity / value</text>\n'
    s += f'  <text x="{x0+w-210}" y="{y0+25}" font-family="Arial" font-size="12" fill="{GRAY}">Premium + specialty moat</text>\n'
    for label, x, y, color in points:
        s += f'  <circle cx="{xp(x):.1f}" cy="{yp(y):.1f}" r="10" fill="{color}" stroke="{DARK}" stroke-width="1"/>\n'
        s += f'  <text x="{xp(x)+14:.1f}" y="{yp(y)+4:.1f}" font-family="Arial" font-size="12" font-weight="{"700" if label=="Michelin" else "400"}" fill="{color if label=="Michelin" else DARK}">{escape(label)}</text>\n'
    s += f'  <text x="120" y="552" font-family="Arial" font-size="13" fill="{DARK}">Use this as a visual explanation of Michelin\'s moat: premium brand, R&D, homologations, distribution and specialty tire expertise.</text>\n'
    s += footer("Source: qualitative positioning based on company disclosures and peer descriptions.")
    return s


def drivers_risks_heatmap() -> str:
    rows = [
        ("Driver", "Premiumisation", 5, 4),
        ("Driver", "18-inch+ tires", 5, 4),
        ("Driver", "EV tire demand", 4, 4),
        ("Driver", "Connected fleets", 3, 3),
        ("Driver", "Specialties", 4, 3),
        ("Risk", "Low-cost imports", 4, 4),
        ("Risk", "FX / EUR strength", 4, 3),
        ("Risk", "Raw materials", 4, 3),
        ("Risk", "OE weakness", 5, 3),
        ("Risk", "Factory under-utilization", 4, 3),
    ]
    s = header(
        "Drivers and risks heatmap",
        "Qualitative scoring from 1 to 5. The goal is to show which items deserve most attention in the thesis.",
    )
    x0, y0 = 90, 105
    cols = [70, 330, 560, 720]
    s += f'  <rect x="{x0}" y="{y0}" width="850" height="38" fill="{BLUE}"/>\n'
    for text, x in [("Type", cols[0]), ("Item", cols[1]), ("Impact", cols[2]), ("Likelihood", cols[3])]:
        s += f'  <text x="{x}" y="{y0+25}" font-family="Arial" font-size="13" font-weight="700" fill="#ffffff">{text}</text>\n'
    def col(score):
        return {5: RED, 4: GOLD, 3: TEAL, 2: "#A7C7E7", 1: "#DDE7EE"}[score]
    for i, (typ, item, impact, likelihood) in enumerate(rows):
        y = y0 + 38 + i * 38
        bg = "#FFFFFF" if i % 2 == 0 else LIGHT
        s += f'  <rect x="{x0}" y="{y}" width="850" height="38" fill="{bg}" stroke="{MID}"/>\n'
        typ_color = TEAL if typ == "Driver" else RED
        s += f'  <text x="{cols[0]}" y="{y+24}" font-family="Arial" font-size="12" font-weight="700" fill="{typ_color}">{typ}</text>\n'
        s += f'  <text x="{cols[1]}" y="{y+24}" font-family="Arial" font-size="12" fill="{DARK}">{escape(item)}</text>\n'
        for score, x in [(impact, cols[2]), (likelihood, cols[3])]:
            s += f'  <rect x="{x}" y="{y+7}" width="70" height="24" rx="4" fill="{col(score)}"/>\n'
            s += f'  <text x="{x+35}" y="{y+24}" text-anchor="middle" font-family="Arial" font-size="12" font-weight="700" fill="#ffffff">{score}/5</text>\n'
    s += f'  <text x="90" y="548" font-family="Arial" font-size="13" fill="{DARK}">Readout: premiumization and large-rim/EV tires are high-impact drivers; low-cost imports, FX and OE weakness are the risks to monitor.</text>\n'
    s += footer("Source: qualitative assessment from Michelin results commentary and industry logic.")
    return s


def segment_mix_margin() -> str:
    data = [
        ("Auto", 55.0, 11.7, TEAL),
        ("Road transport", 23.2, 4.7, GOLD),
        ("Specialties", 21.8, 13.5, BLUE),
    ]
    s = header(
        "Michelin 2025 segment mix and profitability",
        "Sales mix by reporting segment and EBIT margin. Auto is largest; specialties are smaller but high-margin.",
    )
    x0, y0, w, h = 120, 145, 760, 56
    cur = x0
    for label, share, margin, color in data:
        bw = w * share / 100
        s += f'  <rect x="{cur:.1f}" y="{y0}" width="{bw:.1f}" height="{h}" fill="{color}"/>\n'
        s += f'  <text x="{cur+bw/2:.1f}" y="{y0+34}" text-anchor="middle" font-family="Arial" font-size="13" font-weight="700" fill="#ffffff">{share:.1f}%</text>\n'
        cur += bw
    s += f'  <rect x="{x0}" y="{y0}" width="{w}" height="{h}" fill="none" stroke="{DARK}"/>\n'
    s += '  <g font-family="Arial, Helvetica, sans-serif" font-size="13" fill="#20252B">\n'
    for i, (label, share, margin, color) in enumerate(data):
        y = 260 + i * 72
        s += f'    <rect x="130" y="{y-18}" width="18" height="18" fill="{color}"/>\n'
        s += f'    <text x="160" y="{y-4}" font-weight="700">{escape(label)}</text>\n'
        s += f'    <text x="330" y="{y-4}">Sales mix: {share:.1f}%</text>\n'
        s += f'    <rect x="500" y="{y-28}" width="{margin*26:.1f}" height="24" fill="{color}" opacity="0.85"/>\n'
        s += f'    <text x="{510+margin*26:.1f}" y="{y-10}">EBIT margin {margin:.1f}%</text>\n'
    s += "  </g>\n"
    s += f'  <text x="120" y="535" font-family="Arial" font-size="13" fill="{DARK}">Readout: the mix matters. Road transport has the weakest 2025 margin, while Auto and Specialties carry most of the profitability.</text>\n'
    s += footer("Source: Michelin key figures 2020-2025, Sales and Reporting segments tabs.")
    return s


def historical_pnl_combo() -> str:
    years = [2020, 2021, 2022, 2023, 2024, 2025]
    sales = [20469, 23795, 28590, 28343, 27193, 25992]
    margins = [9.2, 12.5, 11.9, 12.6, 12.4, 10.5]
    s = header("Historical sales and EBIT margin", "Bars show revenue; line shows segment EBIT margin.")
    x0, y0, w, h = 115, 105, 780, 360
    max_sales = 30000
    s += f'  <line x1="{x0}" y1="{y0+h}" x2="{x0+w}" y2="{y0+h}" stroke="{DARK}"/>\n'
    s += f'  <line x1="{x0}" y1="{y0}" x2="{x0}" y2="{y0+h}" stroke="{DARK}"/>\n'
    bar_gap = w / len(years)
    pts = []
    for i, (yr, sale, margin) in enumerate(zip(years, sales, margins)):
        bx = x0 + i * bar_gap + 28
        bh = h * sale / max_sales
        by = y0 + h - bh
        s += f'  <rect x="{bx:.1f}" y="{by:.1f}" width="60" height="{bh:.1f}" fill="{TEAL}"/>\n'
        s += f'  <text x="{bx+30:.1f}" y="{y0+h+24}" text-anchor="middle" font-family="Arial" font-size="11">{yr}</text>\n'
        s += f'  <text x="{bx+30:.1f}" y="{by-8:.1f}" text-anchor="middle" font-family="Arial" font-size="10" fill="{DARK}">{sale/1000:.1f}</text>\n'
        px = bx + 30
        py = y0 + h - (margin - 8) / 6 * h
        pts.append((px, py, margin))
    for (x1, y1, _), (x2, y2, _) in zip(pts, pts[1:]):
        s += f'  <line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" stroke="{RED}" stroke-width="3"/>\n'
    for x, y, m in pts:
        s += f'  <circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="{RED}"/>\n'
        s += f'  <text x="{x+8:.1f}" y="{y-8:.1f}" font-family="Arial" font-size="10" fill="{RED}">{m:.1f}%</text>\n'
    s += f'  <text x="115" y="510" font-family="Arial" font-size="13" fill="{DARK}">Readout: 2025 sales are below the 2022 peak, but margins remain structurally above 2020; the issue is volume/absorption, not a collapse in pricing.</text>\n'
    s += footer("Source: Michelin key figures 2020-2025.")
    return s


def forecast_pnl_combo() -> str:
    years = ["2025A", "2026E", "2027E", "2028E", "2029E", "2030E"]
    sales = [25.992, 26.534, 27.197, 28.013, 28.854, 29.575]
    net_income = [1.664, 1.866, 1.940, 2.033, 2.130, 2.221]
    net_margin = [ni / s * 100 for ni, s in zip(net_income, sales)]
    s = header(
        "Projection du CA, du resultat net et de la marge nette",
        "2026E ancre sur le consensus Michelin; 2027E-2030E derives du scenario central prudent.",
        width=1120,
        height=640,
    )
    x0, y0, w, h = 120, 125, 820, 325
    revenue_max = 32.0
    pnl_min, pnl_max = 1.4, 3.4

    # Gridlines and axes
    for tick in [0, 8, 16, 24, 32]:
        y = y0 + h - (tick / revenue_max) * h
        s += f'  <line x1="{x0}" y1="{y:.1f}" x2="{x0+w}" y2="{y:.1f}" stroke="{MID}" stroke-width="1"/>\n'
        s += f'  <text x="{x0-12}" y="{y+4:.1f}" text-anchor="end" font-family="Arial" font-size="11" fill="{GRAY}">{tick:.0f}</text>\n'
    for tick in [1.5, 2.0, 2.5, 3.0, 3.5]:
        y = y0 + h - ((tick - pnl_min) / (pnl_max - pnl_min)) * h
        s += f'  <text x="{x0+w+12}" y="{y+4:.1f}" font-family="Arial" font-size="11" fill="{GRAY}">{tick:.1f}</text>\n'
    s += f'  <line x1="{x0}" y1="{y0+h}" x2="{x0+w}" y2="{y0+h}" stroke="{DARK}" stroke-width="1.5"/>\n'
    s += f'  <line x1="{x0}" y1="{y0}" x2="{x0}" y2="{y0+h}" stroke="{DARK}" stroke-width="1.5"/>\n'
    s += f'  <line x1="{x0+w}" y1="{y0}" x2="{x0+w}" y2="{y0+h}" stroke="{DARK}" stroke-width="1.5"/>\n'
    s += f'  <text x="{x0}" y="{y0-18}" font-family="Arial" font-size="12" font-weight="700" fill="{DARK}">CA (EUR Md)</text>\n'
    s += f'  <text x="{x0+w-5}" y="{y0-18}" text-anchor="end" font-family="Arial" font-size="12" font-weight="700" fill="{DARK}">Resultat net (EUR Md)</text>\n'

    gap = w / len(years)
    bar_w = 74
    net_points = []
    for i, (year, sale, net_inc, margin) in enumerate(zip(years, sales, net_income, net_margin)):
        bx = x0 + i * gap + (gap - bar_w) / 2
        bh = h * (sale / revenue_max)
        by = y0 + h - bh
        color = GRAY if i == 0 else BLUE
        s += f'  <rect x="{bx:.1f}" y="{by:.1f}" width="{bar_w}" height="{bh:.1f}" fill="{color}" opacity="0.96"/>\n'
        s += f'  <text x="{bx + bar_w/2:.1f}" y="{y0+h+24}" text-anchor="middle" font-family="Arial" font-size="11">{escape(year)}</text>\n'
        s += f'  <text x="{bx + bar_w/2:.1f}" y="{by-8:.1f}" text-anchor="middle" font-family="Arial" font-size="10" fill="{DARK}">{sale:.1f}</text>\n'
        px = bx + bar_w / 2
        net_y = y0 + h - ((net_inc - pnl_min) / (pnl_max - pnl_min)) * h
        net_points.append((px, net_y, net_inc, margin))

    separator_x = x0 + gap
    s += f'  <line x1="{separator_x:.1f}" y1="{y0-5}" x2="{separator_x:.1f}" y2="{y0+h+6}" stroke="{GRAY}" stroke-width="1.5" stroke-dasharray="5 4"/>\n'

    for (x1, y1, _, _), (x2, y2, _, _) in zip(net_points, net_points[1:]):
        s += f'  <line x1="{x1:.1f}" y1="{y1:.1f}" x2="{x2:.1f}" y2="{y2:.1f}" stroke="{GOLD}" stroke-width="3"/>\n'
    for i, (x, y, value, margin) in enumerate(net_points):
        s += f'  <circle cx="{x:.1f}" cy="{y:.1f}" r="5.5" fill="{GOLD}" stroke="#ffffff" stroke-width="1.5"/>\n'
        s += f'  <text x="{x+8:.1f}" y="{y-8:.1f}" font-family="Arial" font-size="10" fill="{GOLD}">{value:.1f}</text>\n'
        margin_dx = 10 if i < len(net_points) - 1 else 12
        margin_dy = 18 if i % 2 == 0 else 22
        s += f'  <text x="{x+margin_dx:.1f}" y="{y+margin_dy:.1f}" font-family="Arial" font-size="10" fill="{TEAL}">{margin:.1f}%</text>\n'

    # Legend below chart to avoid collisions with axis titles.
    lx, ly = 220, 498
    s += f'  <rect x="{lx}" y="{ly}" width="14" height="14" fill="{GRAY}"/>\n'
    s += f'  <text x="{lx+22}" y="{ly+12}" font-family="Arial" font-size="11" fill="{DARK}">2025A publie</text>\n'
    s += f'  <rect x="{lx+135}" y="{ly}" width="14" height="14" fill="{BLUE}"/>\n'
    s += f'  <text x="{lx+157}" y="{ly+12}" font-family="Arial" font-size="11" fill="{DARK}">CA 2026E-2030E</text>\n'
    s += f'  <line x1="{lx+325}" y1="{ly+7}" x2="{lx+345}" y2="{ly+7}" stroke="{GOLD}" stroke-width="3"/>\n'
    s += f'  <text x="{lx+353}" y="{ly+12}" font-family="Arial" font-size="11" fill="{DARK}">Resultat net</text>\n'
    s += f'  <text x="{lx+498}" y="{ly+12}" font-family="Arial" font-size="11" fill="{TEAL}">Marge nette en label</text>\n'

    s += f"  <text x=\"120\" y=\"545\" font-family=\"Arial\" font-size=\"12\" fill=\"{DARK}\">Lecture: la hausse du resultat net vient a la fois de la progression du CA et de l'amelioration graduelle de la marge nette.</text>\n"
    s += f"  <text x=\"120\" y=\"564\" font-family=\"Arial\" font-size=\"12\" fill=\"{DARK}\">2026E reprend le consensus Michelin; 2027E-2030E prolongent ce point d'ancrage dans le scenario central.</text>\n"
    s += footer("Sources: Michelin, resultats annuels 2025; Michelin consensus du 9 mars 2026; scenario central interne pour 2027E-2030E.", y=612)
    return s


def peer_screening_matrix() -> str:
    cols = ["Tire exposure", "Premium mix", "Geography", "Leverage risk", "Comparable?"]
    rows = [
        ("Bridgestone", [5, 4, 5, 4, 5]),
        ("Goodyear", [5, 2, 4, 1, 3]),
        ("Continental", [2, 3, 4, 4, 3]),
        ("Pirelli", [5, 5, 3, 3, 5]),
        ("Yokohama", [4, 3, 3, 4, 4]),
        ("Hankook", [5, 3, 3, 4, 4]),
    ]
    s = header("Peer screening matrix", "Qualitative comparability screen. 5 = strongest / most comparable.")
    x0, y0, cellw, cellh = 90, 120, 135, 46
    s += f'  <rect x="{x0}" y="{y0}" width="{cellw*6}" height="{cellh}" fill="{BLUE}"/>\n'
    s += f'  <text x="{x0+12}" y="{y0+29}" font-family="Arial" font-size="12" font-weight="700" fill="#ffffff">Peer</text>\n'
    for i, col in enumerate(cols):
        s += f'  <text x="{x0+cellw*(i+1)+10}" y="{y0+29}" font-family="Arial" font-size="11" font-weight="700" fill="#ffffff">{escape(col)}</text>\n'
    def color(score):
        return {5: TEAL, 4: "#73BFAE", 3: GOLD, 2: "#E8A0AA", 1: RED}[score]
    for r, (name, scores) in enumerate(rows):
        y = y0 + cellh * (r + 1)
        s += f'  <rect x="{x0}" y="{y}" width="{cellw}" height="{cellh}" fill="{LIGHT}" stroke="{MID}"/>\n'
        s += f'  <text x="{x0+12}" y="{y+29}" font-family="Arial" font-size="12" font-weight="700" fill="{DARK}">{escape(name)}</text>\n'
        for c, score in enumerate(scores):
            x = x0 + cellw * (c + 1)
            s += f'  <rect x="{x}" y="{y}" width="{cellw}" height="{cellh}" fill="{color(score)}" stroke="#ffffff"/>\n'
            s += f'  <text x="{x+cellw/2}" y="{y+29}" text-anchor="middle" font-family="Arial" font-size="12" font-weight="700" fill="#ffffff">{score}/5</text>\n'
    s += f'  <text x="90" y="520" font-family="Arial" font-size="13" fill="{DARK}">Readout: Pirelli and Bridgestone are the cleanest quality peers; Goodyear is relevant sectorally but less clean because leverage and weak margins distort multiples.</text>\n'
    s += footer("Source: qualitative assessment based on business mix and StockAnalysis valuation data.")
    return s


def wacc_build_up() -> str:
    s = header("WACC build-up", "The discount rate is built from cost of equity, after-tax cost of debt and market capital structure.")
    cards = [
        ("Risk-free rate", "3.7%", BLUE),
        ("Equity risk premium", "5.8%", TEAL),
        ("Beta", "1.00x", GOLD),
        ("Cost of equity", "9.4%", RED),
        ("After-tax cost of debt", "2.8%", BLUE),
        ("Equity / debt weights", "90.5% / 9.5%", TEAL),
        ("WACC", "8.8%", RED),
    ]
    x, y = 90, 125
    for i, (label, value, color) in enumerate(cards):
        cx = x + (i % 4) * 230
        cy = y + (i // 4) * 145
        s += f'  <rect x="{cx}" y="{cy}" width="195" height="95" rx="8" fill="{LIGHT}" stroke="{MID}"/>\n'
        s += f'  <text x="{cx+97}" y="{cy+38}" text-anchor="middle" font-family="Arial" font-size="22" font-weight="700" fill="{color}">{escape(value)}</text>\n'
        s += f'  <text x="{cx+97}" y="{cy+66}" text-anchor="middle" font-family="Arial" font-size="12" fill="{DARK}">{escape(label)}</text>\n'
    s += f'  <text x="90" y="430" font-family="Arial" font-size="14" font-weight="700" fill="{BLUE}">Formula logic</text>\n'
    s += f'  <text x="90" y="460" font-family="Arial" font-size="13" fill="{DARK}">Cost of equity = risk-free rate + beta x equity risk premium. WACC then blends cost of equity and after-tax cost of debt using market value weights.</text>\n'
    s += f'  <text x="90" y="490" font-family="Arial" font-size="13" fill="{DARK}">This makes the DCF discount rate transparent: the main driver is equity cost because Michelin is mostly equity-financed in market value terms.</text>\n'
    s += footer("Source: project WACC model; CountryEconomy, Kroll, StockAnalysis.")
    return s


def football_field_summary() -> str:
    scenario_payload = scenario_model_values()
    dcf_low = next(item["price"] for item in scenario_payload["scenarios"] if item["name"] == "Bear")
    dcf_mid = next(item["price"] for item in scenario_payload["scenarios"] if item["name"] == "Base")
    dcf_high = next(item["price"] for item in scenario_payload["scenarios"] if item["name"] == "Bull")
    comps_low, comps_mid, comps_high = 23.5, 36.1, 44.1
    target = (comps_mid + dcf_mid) / 2
    ranges = [
        ("Cours actuel", 32.2, 32.2, 32.2, GRAY),
        ("Comparables", comps_low, comps_mid, comps_high, TEAL),
        ("DCF scenarios", dcf_low, dcf_mid, dcf_high, BLUE),
        ("Objectif retenu", target, target, target, RED),
    ]
    minv, maxv = 20, 65
    x0, y0, w = 230, 125, 660
    def xp(v): return x0 + (v - minv) / (maxv - minv) * w
    s = header("Football field summary", "Valuation ranges by method, EUR/share. DCF range now reflects Bear / Base / Bull scenarios.")
    for tick in [20, 30, 40, 50, 60]:
        x = xp(tick)
        s += f'  <line x1="{x:.1f}" y1="95" x2="{x:.1f}" y2="430" stroke="{MID}"/>\n'
        s += f'  <text x="{x-8:.1f}" y="455" font-family="Arial" font-size="11" fill="{GRAY}">{tick}</text>\n'
    for i, (label, low, mid, high, color) in enumerate(ranges):
        y = y0 + i * 78
        s += f'  <text x="85" y="{y+8}" font-family="Arial" font-size="13" font-weight="700" fill="{DARK}">{escape(label)}</text>\n'
        s += f'  <rect x="{xp(low):.1f}" y="{y-8}" width="{max(3, xp(high)-xp(low)):.1f}" height="20" fill="{color}" opacity="0.85"/>\n'
        s += f'  <circle cx="{xp(mid):.1f}" cy="{y+2}" r="7" fill="{DARK}"/>\n'
        s += f'  <text x="{xp(high)+10:.1f}" y="{y+7}" font-family="Arial" font-size="11" fill="{GRAY}">{low:.1f} - {high:.1f}</text>\n'
    s += f'  <text x="90" y="515" font-family="Arial" font-size="13" fill="{DARK}">Readout: the target price sits between peer valuation and DCF, while the DCF range now reflects explicit Bear / Base / Bull scenarios rather than only WACC/g sensitivity.</text>\n'
    s += footer("Source: Michelin valuation model workbook; comparables summary table.")
    return s


def investment_scorecard() -> str:
    rows = [
        ("Valuation upside", "Green", "+24.9% price upside to blended target"),
        ("Cash-flow generation", "Green", "FCF EUR 2.2bn in 2025 despite weak volumes"),
        ("Balance sheet", "Green", "Net debt / EBITDA around 0.5x"),
        ("Margin recovery", "Yellow", "Key thesis driver; needs volume normalization"),
        ("FX and low-cost imports", "Yellow", "Main risks to monitor"),
    ]
    s = header("Investment thesis scorecard", "Simple visual summary of why the recommendation is Buy, with risks kept visible.")
    x0, y0 = 90, 115
    s += f'  <rect x="{x0}" y="{y0}" width="860" height="40" fill="{BLUE}"/>\n'
    for x, t in [(110, "Factor"), (330, "Signal"), (500, "Comment")]:
        s += f'  <text x="{x}" y="{y0+26}" font-family="Arial" font-size="13" font-weight="700" fill="#ffffff">{t}</text>\n'
    for i, (factor, signal, comment) in enumerate(rows):
        y = y0 + 40 + i * 65
        bg = "#ffffff" if i % 2 == 0 else LIGHT
        color = TEAL if signal == "Green" else GOLD
        s += f'  <rect x="{x0}" y="{y}" width="860" height="65" fill="{bg}" stroke="{MID}"/>\n'
        s += f'  <text x="110" y="{y+38}" font-family="Arial" font-size="13" font-weight="700" fill="{DARK}">{escape(factor)}</text>\n'
        s += f'  <circle cx="365" cy="{y+32}" r="13" fill="{color}"/>\n'
        s += f'  <text x="500" y="{y+38}" font-family="Arial" font-size="13" fill="{DARK}">{escape(comment)}</text>\n'
    s += f'  <text x="90" y="520" font-family="Arial" font-size="13" fill="{DARK}">Readout: the thesis is attractive but not risk-free; the scorecard keeps both upside and execution risks visible.</text>\n'
    s += footer("Source: project valuation model and Michelin 2025 results.")
    return s


def main() -> None:
    charts = {
        "market_segmentation.svg": market_segmentation(),
        "competitive_positioning.svg": competitive_positioning(),
        "segment_mix_margin.svg": segment_mix_margin(),
        "historical_pnl_combo.svg": historical_pnl_combo(),
        "forecast_pnl_combo.svg": forecast_pnl_combo(),
        "scenario_dcf_values.svg": scenario_dcf_values(),
        "tornado_sensitivity.svg": tornado_sensitivity(),
        "buyback_impact.svg": buyback_impact(),
        "football_field_summary.svg": football_field_summary(),
    }
    for name, svg in charts.items():
        save_svg(name, svg)
        print(OUT / name)
    for name in LEGACY_REMOVED_CHARTS:
        path = OUT / name
        if path.exists():
            path.unlink()
            print(f"removed {path}")


if __name__ == "__main__":
    main()
